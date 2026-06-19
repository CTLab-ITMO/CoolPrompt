"""FastAPI app for CoolPrompt Interface Demo."""

from __future__ import annotations

import logging
import time
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from csv import Sniffer, reader
from io import BytesIO, StringIO
from pathlib import Path
from threading import Lock
from typing import Any

from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from .methods import METHOD_BY_ID, public_methods
from .runner import run_single_optimization
from .schemas import JobCreateRequest, JobStatus
from .settings import get_settings


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

settings = get_settings()
executor = ThreadPoolExecutor(max_workers=settings.max_workers)
jobs: dict[str, JobStatus] = {}
jobs_lock = Lock()
logger = logging.getLogger("demo_service")
logger.setLevel(logging.INFO)

app = FastAPI(title=settings.app_name, version="0.1.0")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


MAX_IMPORT_ROWS = 500


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _first_two_values(row: list[Any] | tuple[Any, ...]) -> tuple[str, str] | None:
    values = [_cell_text(cell) for cell in row]
    non_empty = [value for value in values if value]
    if len(non_empty) < 2:
        return None
    return non_empty[0], non_empty[1]


def _looks_like_header(first: str, second: str) -> bool:
    normalized = {first.strip().lower(), second.strip().lower()}
    input_headers = {
        "input",
        "source",
        "text",
        "sample",
        "prompt",
        "вход",
        "пример",
        "текст",
        "запрос",
        "сообщение",
    }
    target_headers = {
        "target",
        "label",
        "output",
        "answer",
        "expected",
        "метка",
        "эталон",
        "ответ",
        "класс",
        "результат",
    }
    return bool(normalized & input_headers) and bool(normalized & target_headers)


def _normalize_import_rows(rows: list[tuple[str, str]]) -> list[dict[str, str]]:
    if rows and _looks_like_header(rows[0][0], rows[0][1]):
        rows = rows[1:]
    normalized = [
        {"input": input_text, "target": target_text}
        for input_text, target_text in rows
        if input_text and target_text
    ]
    return normalized[:MAX_IMPORT_ROWS]


def _decode_tabular_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _parse_csv_samples(content: bytes) -> list[dict[str, str]]:
    text = _decode_tabular_text(content)
    sample = text[:4096]
    try:
        dialect = Sniffer().sniff(sample, delimiters=",;\t")
    except Exception:
        dialect = "excel"
    parsed_rows: list[tuple[str, str]] = []
    for row in reader(StringIO(text), dialect):
        pair = _first_two_values(row)
        if pair:
            parsed_rows.append(pair)
    return _normalize_import_rows(parsed_rows)


def _parse_xlsx_samples(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    parsed_rows: list[tuple[str, str]] = []
    for row in sheet.iter_rows(values_only=True):
        pair = _first_two_values(row)
        if pair:
            parsed_rows.append(pair)
        if len(parsed_rows) > MAX_IMPORT_ROWS + 1:
            break
    workbook.close()
    return _normalize_import_rows(parsed_rows)


def _parse_sample_file(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx_samples(content)
    if suffix in {".csv", ".tsv", ".txt"}:
        return _parse_csv_samples(content)
    raise HTTPException(status_code=400, detail="Поддерживаются только CSV, TSV и XLSX-файлы.")


def _public_error_message(exc: Exception) -> str:
    """Return a short customer-facing error instead of provider JSON dumps."""

    message = str(exc)
    lowered = message.lower()
    if "content_filter" in lowered or "content management policy" in lowered:
        return (
            "Провайдер модели отклонил запрос своим фильтром безопасности. "
            "Попробуйте Gemini 2.5 Flash или DeepSeek Chat v3, либо немного измените промпт."
        )
    if "unsupported_country_region_territory" in lowered or "request_forbidden" in lowered:
        return (
            "Провайдер модели недоступен из текущего региона. "
            "Выберите Gemini 2.5 Flash или DeepSeek Chat v3."
        )
    if "invalid_api_key" in lowered or "incorrect api key" in lowered:
        return "Ключ модели не принят провайдером. Проверьте переменные окружения сервиса."
    return message[:600]


def _model_options() -> list[dict[str, str]]:
    base_url = (settings.openai_base_url or "").lower()
    if "openrouter" in base_url:
        return [
            {"value": "openai/gpt-4.1-mini", "label": "OpenAI GPT-4.1 mini"},
            {"value": "openai/gpt-4o-mini", "label": "OpenAI GPT-4o mini"},
            {"value": "google/gemini-2.5-flash", "label": "Gemini 2.5 Flash"},
            {"value": "deepseek/deepseek-chat-v3-0324", "label": "DeepSeek Chat v3"},
        ]
    return [
        {"value": "gpt-4o-mini", "label": "gpt-4o-mini"},
        {"value": "gpt-4o", "label": "gpt-4o"},
        {"value": "gpt-4.1-mini", "label": "gpt-4.1-mini"},
        {"value": "gpt-4.1-nano", "label": "gpt-4.1-nano"},
    ]


@app.middleware("http")
async def add_no_cache_headers(request: Request, call_next):
    response = await call_next(request)
    if request.url.path == "/" or request.url.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, max-age=0"
    return response


def _store(job: JobStatus) -> None:
    with jobs_lock:
        jobs[job.job_id] = job


def _patch_job(job_id: str, **updates: Any) -> None:
    with jobs_lock:
        job = jobs[job_id]
        if job.status in {"completed", "failed"}:
            return
        data = job.model_dump()
        data.update(updates)
        data["updated_at"] = time.time()
        jobs[job_id] = JobStatus(**data)


def _run_job(job_id: str, payload: JobCreateRequest) -> None:
    def progress(stage: str, percent: int, message: str) -> None:
        logger.info("job=%s stage=%s percent=%s message=%s", job_id, stage, percent, message)
        _patch_job(
            job_id,
            status="running",
            progress_stage=stage,
            progress_percent=percent,
            progress_message=message,
        )

    progress("preparing", 10, "Готовим запуск")
    try:
        logger.info("job=%s mode=%s started", job_id, payload.mode)
        assert payload.request is not None
        result = run_single_optimization(payload.request, settings, progress_callback=progress)
        _patch_job(
            job_id,
            status="completed",
            progress_stage="completed",
            progress_percent=100,
            progress_message="Оптимизация завершена",
            result=result,
        )
        logger.info("job=%s completed", job_id)
    except Exception as exc:  # noqa: BLE001 - surfaced to UI as job error
        logger.exception("job=%s failed", job_id)
        _patch_job(
            job_id,
            status="failed",
            progress_stage="failed",
            progress_percent=100,
            progress_message="Оптимизация завершилась с ошибкой",
            error=_public_error_message(exc),
        )


@app.get("/")
def index() -> FileResponse:
    """Serve the one-page demo interface."""

    return FileResponse(STATIC_DIR / "index.html")


@app.head("/")
def head_index() -> Response:
    """Render probes the root URL with HEAD before declaring the service live."""

    return Response(status_code=200)


@app.get("/api/health")
def health() -> dict[str, str]:
    """Railway healthcheck endpoint."""

    return {"status": "ok"}


@app.get("/api/config")
def config() -> dict[str, Any]:
    """Expose non-secret frontend configuration."""

    return {
        "appName": settings.app_name,
        "defaultModel": settings.model_name,
        "hasOpenAIKey": settings.has_openai_key,
        "allowMock": settings.allow_mock,
        "forceMock": settings.force_mock,
        "modelOptions": _model_options(),
    }


@app.get("/api/methods")
def methods() -> list[dict[str, Any]]:
    """Return method catalog for the UI."""

    return public_methods()


@app.post("/api/samples/import")
async def import_samples(file: UploadFile = File(...)) -> dict[str, Any]:
    """Parse two-column CSV/TSV/XLSX samples for the demo dataset editor."""

    filename = file.filename or "samples"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой.")

    try:
        rows = _parse_sample_file(filename, content)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001 - converted to a customer-facing error
        logger.exception("sample import failed filename=%s", filename)
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}") from exc

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Не удалось найти две заполненные колонки с примерами и эталонами.",
        )
    return {"filename": filename, "count": len(rows), "rows": rows}


@app.post("/api/jobs", response_model=JobStatus)
def create_job(payload: JobCreateRequest) -> JobStatus:
    """Create a background optimization job."""

    if payload.request is not None and payload.request.method not in METHOD_BY_ID:
        raise HTTPException(status_code=400, detail=f"Unknown method: {payload.request.method}")

    job = JobStatus(
        job_id=uuid.uuid4().hex,
        mode=payload.mode,
        status="queued",
        created_at=time.time(),
        updated_at=time.time(),
        progress_stage="queued",
        progress_percent=5,
        progress_message="Задача ожидает запуска",
    )
    _store(job)
    assert payload.request is not None
    logger.info(
        "job=%s queued mode=single method=%s model=%s mock=%s dataset_size=%s",
        job.job_id,
        payload.request.method,
        payload.request.model_name or settings.model_name,
        payload.request.mock or settings.force_mock,
        len(payload.request.dataset or []),
    )
    future: Future = executor.submit(_run_job, job.job_id, payload)
    future.add_done_callback(lambda _: None)
    return job


@app.get("/api/jobs/{job_id}", response_model=JobStatus)
def get_job(job_id: str) -> JobStatus:
    """Read the current job state."""

    with jobs_lock:
        job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return job
